import pandas as pd
import openpyxl
import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import logging
import time

# Configure logging
logging.basicConfig(filename='email.log', level=logging.INFO, format='%(asctime)s - %(levelname)s: %(message)s')

# Function to send email
def send_email(to_email, subject, body):
    # Email configuration
    from_email = 'YOUR_EMAIL@example.com'
    password = 'YOUR_EMAIL_PASSWORD'

    msg = MIMEMultipart()
    msg['From'] = from_email
    msg['To'] = to_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'html'))

    try:
        # Connect to SMTP server and send email
        server = smtplib.SMTP('smtp.example.com', 587)
        server.starttls()
        server.login(from_email, password)
        server.sendmail(from_email, to_email, msg.as_string())
        server.quit()
        
        # Log and print success message
        logging.info(f"Email sent successfully to: {to_email}")
        print(f"Email sent successfully to: {to_email}")
        return True
    except Exception as e:
        # Log and print error message
        logging.error(f"Failed to send email to {to_email}: {e}")
        print(f"Failed to send email to {to_email}: {e}")
        return False

# Read existing data from Excel file
try:
    df_existing = pd.read_excel('hotel_nova.xlsx', parse_dates=['DATUM ISTEKA'])
    logging.info("Existing data loaded successfully.")
except FileNotFoundError:
    df_existing = pd.DataFrame()
    logging.info("No existing data found.")

# Main loop to check for notifications
while True:
    # Calculate today's date and time
    today_date = datetime.datetime.now().date()
    today_time = datetime.datetime.now().strftime('%H:%M:%S')
    logging.info(f"Today's date: {today_date}")

    # Iterate through rows
    for index, row in df_existing.iterrows():
        company_name = row['TVRTKA']
        expiry_date = row['DATUM ISTEKA']
        email_sent = row['SLANJE ISTEKA']  # Assuming 'SLANJE ISTEKA' column contains 'POSLANO' or is empty

        # Convert expiry_date to datetime object if it's not already 'POSLANO'
        if email_sent != 'POSLANO':
            try:
                expiry_date = pd.to_datetime(expiry_date, format='%d/%m/%Y')
            except Exception as e:
                logging.error(f"Error converting date for company {company_name}: {e}")
                continue

            # Check if expiry date is valid
            if pd.notnull(expiry_date):
                # Calculate days until expiration
                days_until_expiry = (expiry_date.date() - today_date).days

                # Define notification intervals
                notification_intervals = [-60, -30, -1, 30, 60]

                # Iterate over notification intervals
                for interval in notification_intervals:
                    # Calculate notification date
                    notification_date = expiry_date + datetime.timedelta(days=interval)

                    # Check if notification date is today
                    if notification_date.date() == today_date:
                        to_email = row['EMAIL']
                        card_number = row['BROJ KARTICE']
                        subject = f'YOUR_SUBJECT_HERE ({card_number})'
                        body = f"""<p style="font-family: Calibri;">YOUR_BODY_HERE</p>"""

                        if send_email(to_email, subject, body):
                            print(f"Updating columns for {company_name}")
                            # Update 'SLANJE ISTEKA' and 'VRIJEME POSLANO' columns to mark email as sent
                            try:
                                df_existing.at[index, 'SLANJE ISTEKA'] = 'POSLANO'
                                df_existing.at[index, 'VRIJEME POSLANO'] = str(today_time)
                                df_existing.at[index, 'DATUM POSLANO'] = today_date.strftime('%d/%m/%Y')
                            except Exception as e:
                                logging.error(f"Error updating columns for company {company_name}: {e}")
                                continue

                            # Remove values in 'DATUM POSLANO' and 'VRIJEME POSLANO' if 'DATUM ISTEKA' or 'BROJ KARTICE' are updated manually
                            try:
                                if row['DATUM POSLANO'] and row['VRIJEME POSLANO'] and pd.notnull(row['DATUM POSLANO']) and pd.notnull(row['VRIJEME POSLANO']):
                                    df_existing.at[index, 'DATUM POSLANO'] = None
                                    df_existing.at[index, 'VRIJEME POSLANO'] = None
                            except Exception as e:
                                logging.error(f"Error removing 'DATUM POSLANO' and 'VRIJEME POSLANO' values for company {company_name}: {e}")
                                continue

    # Save changes to the Excel file
    try:
        # Open the Excel file
        wb = openpyxl.load_workbook('hotel_nova.xlsx')
        ws = wb.active

        # Iterate through rows and update Excel cells
        for index, row in df_existing.iterrows():
            ws.cell(row=index + 2, column=df_existing.columns.get_loc('SLANJE ISTEKA') + 1).value = row['SLANJE ISTEKA']
            ws.cell(row=index + 2, column=df_existing.columns.get_loc('VRIJEME POSLANO') + 1).value = row['VRIJEME POSLANO']
            ws.cell(row=index + 2, column=df_existing.columns.get_loc('DATUM POSLANO') + 1).value = row['DATUM POSLANO']

        # Save changes to the Excel file
        wb.save('hotel_nova.xlsx')
        logging.info("Changes saved to Excel file.")
    except Exception as e:
        logging.error(f"Error saving changes to Excel file: {e}")

    # Wait for 8 hours before the next iteration
    time.sleep(28800)  # 8 hours in seconds
